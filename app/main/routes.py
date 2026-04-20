from datetime import datetime
from collections import defaultdict
import csv
from io import StringIO, BytesIO

from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, make_response, send_file, abort
from flask_login import login_required, current_user

from app.extensions import db
from app.models import RDARecord, TAP, TAPItem, TAP_ENTREGAVEIS, RDAAuditLog


def _registrar_audit(action: str, record: "RDARecord | None", record_id: int = None):
    """Grava uma entrada na pilha de salvamentos do RDA."""
    try:
        nome = current_user.display_name if current_user.is_authenticated else ""
        resumo = ""
        rid = record_id
        if record is not None:
            rid = record.id
            resumo = f"{record.cliente} | {record.realizado} | {record.data}"
        log = RDAAuditLog(
            user_id=current_user.id,
            user_display=nome,
            action=action,
            record_id=rid,
            resumo=resumo,
        )
        db.session.add(log)
        db.session.commit()
    except Exception:
        db.session.rollback()


def _datas_previstas_por_atividade():
    """Mapa {(rotulo_cliente, entregavel): 'YYYY-MM-DD'} usando Fim_Atividade
    do TAP como prazo final. Itens sem fim_atividade ficam fora do mapa."""
    mapa = {}
    for tap in TAP.query.all():
        rotulo = tap.rotulo_cliente
        for item in tap.itens:
            if item.fim_atividade:
                mapa[(rotulo, item.entregavel)] = item.fim_atividade
    return mapa


def _esta_atrasado_por_data(cliente: str, realizado: str, hoje_iso: str,
                            mapa_datas: dict, data_registro: str = None) -> bool:
    """True se existe fim_atividade para (cliente, atividade) e:
    - hoje > fim, OU
    - a data do registro é posterior ao fim (trabalho lançado após o prazo)."""
    dp = mapa_datas.get((cliente or "", realizado or ""))
    if not dp:
        return False
    if dp < hoje_iso:
        return True
    if data_registro and data_registro > dp:
        return True
    return False


def _aplicar_auto_atraso():
    """Aplica regras de auto-atraso cruzando RDA com fim_atividade do TAP.
    Regras (idempotentes):
    (A) Registro Em Andamento cuja data (do registro) > fim_atividade →
        status vira Atrasado + foi_atrasado sticky. Ex.: colaborador lançou
        trabalho no dia 24 para uma atividade com prazo dia 23.
    (B) Registro Concluído cuja data > fim_atividade → apenas marca
        foi_atrasado (sticky), status permanece Concluído.
    (C) Se hoje já passou do fim_atividade e a atividade ainda tem registros
        Em Andamento → os Em Andamento viram Atrasado.
    """
    hoje_iso = datetime.utcnow().strftime("%Y-%m-%d")
    mapa = _datas_previstas_por_atividade()
    total = 0
    for (cliente, atividade), fim in mapa.items():
        if not fim or not cliente or not atividade:
            continue

        # (A) Em Andamento com data do registro > fim → Atrasado + sticky
        n_a = (
            RDARecord.query.filter(
                RDARecord.cliente == cliente,
                RDARecord.realizado == atividade,
                RDARecord.data > fim,
                RDARecord.status_rda == "Em Andamento",
            )
            .update(
                {"status_rda": "Atrasado", "foi_atrasado": True},
                synchronize_session=False,
            )
        )
        total += n_a

        # (B) Concluído com data > fim → só marca sticky (não altera status)
        n_b = (
            RDARecord.query.filter(
                RDARecord.cliente == cliente,
                RDARecord.realizado == atividade,
                RDARecord.data > fim,
                RDARecord.status_rda == "Concluído",
                RDARecord.foi_atrasado == False,
            )
            .update({"foi_atrasado": True}, synchronize_session=False)
        )
        total += n_b

        # (C) Hoje já passou do fim → Em Andamento restante vira Atrasado
        if fim < hoje_iso:
            n_c = (
                RDARecord.query.filter(
                    RDARecord.cliente == cliente,
                    RDARecord.realizado == atividade,
                    RDARecord.status_rda == "Em Andamento",
                )
                .update(
                    {"status_rda": "Atrasado", "foi_atrasado": True},
                    synchronize_session=False,
                )
            )
            total += n_c

    if total:
        db.session.commit()
    return total


main_bp = Blueprint("main", __name__)


def _minutos_intervalo(inicio: str, fim: str) -> int:
    """Retorna duração em minutos. Strings vazias/nulas = 0."""
    if not inicio or not fim:
        return 0
    try:
        i = datetime.strptime(inicio, "%H:%M")
        f = datetime.strptime(fim, "%H:%M")
    except ValueError:
        return 0
    if f < i:
        raise ValueError("A Hora Final não pode ser menor que a Hora Início.")
    return int((f - i).total_seconds() // 60)


def calcular_duracao_total(hi_m: str, hf_m: str, hi_t: str, hf_t: str) -> str:
    total = _minutos_intervalo(hi_m, hf_m) + _minutos_intervalo(hi_t, hf_t)
    return f"{total // 60:02d}:{total % 60:02d}"


def _minutos_registro(r: RDARecord) -> int:
    try:
        h, m = r.duracao.split(":")
        return int(h) * 60 + int(m)
    except (ValueError, AttributeError):
        return 0


def horas_totais(records) -> str:
    total = sum(_minutos_registro(r) for r in records)
    return f"{total // 60:02d}:{total % 60:02d}"


def formatar_data_br(iso: str) -> str:
    """YYYY-MM-DD -> dd/mm/aaaa. Retorna o valor original se não for ISO."""
    if not iso:
        return ""
    try:
        return datetime.strptime(iso, "%Y-%m-%d").strftime("%d/%m/%Y")
    except ValueError:
        return iso


def base_query_for_user():
    if current_user.role in ["gestor", "admin"]:
        return RDARecord.query
    return RDARecord.query.filter_by(user_id=current_user.id)


def aplicar_filtros(query):
    f_data = request.args.get("f_data", "").strip()
    f_cliente = request.args.get("f_cliente", "").strip()
    f_colaborador = request.args.get("f_colaborador", "").strip()
    f_atividade = request.args.get("f_atividade", "").strip()
    f_status = request.args.get("f_status", "").strip()

    if f_data:
        query = query.filter(RDARecord.data == f_data)
    if f_cliente:
        query = query.filter(RDARecord.cliente.ilike(f"%{f_cliente}%"))
    if f_colaborador:
        query = query.filter(RDARecord.colaborador.ilike(f"%{f_colaborador}%"))
    if f_atividade:
        query = query.filter(RDARecord.realizado.ilike(f"%{f_atividade}%"))
    if f_status:
        query = query.filter(RDARecord.status_rda == f_status)

    return query, {
        "f_data": f_data,
        "f_cliente": f_cliente,
        "f_colaborador": f_colaborador,
        "f_atividade": f_atividade,
        "f_status": f_status,
    }


def stats_mes_atual(mes_str: str = None):
    """Retorna dados agregados de um mês específico (YYYY-MM) para a dashboard.
    Se mes_str for None, usa o mês corrente.

    Contagem por atividade (cliente + entregável realizado), não por projeto:
    - Concluído tem prioridade sobre Em Andamento: se a mesma atividade tiver
      algum registro 'Concluído', ela conta apenas em Concluído.
    - Em Andamento = atividades sem nenhum registro Concluído e com pelo
      menos um registro Em Andamento.
    - Atrasado = atividades com pelo menos um registro marcado como Atrasado
      (sticky, permanece contabilizado mesmo após conclusão).
    """
    hoje = datetime.utcnow()
    if mes_str and len(mes_str) == 7 and mes_str[4] == "-":
        try:
            datetime.strptime(mes_str, "%Y-%m")
            prefixo = mes_str
        except ValueError:
            prefixo = hoje.strftime("%Y-%m")
    else:
        prefixo = hoje.strftime("%Y-%m")

    try:
        mes_label = datetime.strptime(prefixo, "%Y-%m").strftime("%m/%Y")
    except ValueError:
        mes_label = hoje.strftime("%m/%Y")

    query = base_query_for_user().filter(RDARecord.data.like(f"{prefixo}-%"))
    records = query.all()

    por_projeto = defaultdict(int)
    for r in records:
        por_projeto[r.cliente or "(sem cliente)"] += _minutos_registro(r)

    # PDI fica fora da contagem de status — é contabilizado à parte (apenas horas).
    mapa_datas = _datas_previstas_por_atividade()
    hoje_iso = datetime.utcnow().strftime("%Y-%m-%d")

    atividades = defaultdict(lambda: {"em_andamento": False, "concluido": False, "foi_atrasado": False})
    for r in records:
        if (r.realizado or "") == "PDI":
            continue
        chave = (r.cliente or "(sem cliente)", r.realizado or "")
        if r.status_rda == "Em Andamento":
            atividades[chave]["em_andamento"] = True
        elif r.status_rda == "Concluído":
            atividades[chave]["concluido"] = True
        if r.foi_atrasado:
            atividades[chave]["foi_atrasado"] = True
        # Auto-atraso: grupo marca atrasado se fim_atividade já passou OU
        # se a data do registro é posterior ao fim (trabalho após o prazo).
        if _esta_atrasado_por_data(
            r.cliente, r.realizado, hoje_iso, mapa_datas, data_registro=r.data
        ):
            if r.status_rda != "Concluído" or r.data > mapa_datas.get((r.cliente or "", r.realizado or ""), ""):
                atividades[chave]["foi_atrasado"] = True

    por_status = {"Em Andamento": 0, "Concluído": 0, "Atrasado": 0}
    for dados in atividades.values():
        if dados["concluido"]:
            por_status["Concluído"] += 1
        elif dados["em_andamento"]:
            por_status["Em Andamento"] += 1
        if dados["foi_atrasado"]:
            por_status["Atrasado"] += 1

    total_min = sum(por_projeto.values())

    pdi_min = sum(_minutos_registro(r) for r in records if (r.realizado or "") == "PDI")

    # Para o gráfico de horas por projeto, remove o sufixo " / CTRL Nº"
    # e deixa somente o nome do cliente.
    labels = [k.split(" / ")[0] for k in por_projeto.keys()]
    horas_decimais = [round(m / 60, 2) for m in por_projeto.values()]

    return {
        "mes_key": prefixo,
        "mes_label": mes_label,
        "total_horas_mes": f"{total_min // 60:02d}:{total_min % 60:02d}",
        "labels_projeto": labels,
        "horas_por_projeto": horas_decimais,
        "total_registros_mes": len(records),
        "por_status": por_status,
        "total_horas_pdi": f"{pdi_min // 60:02d}:{pdi_min % 60:02d}",
        "pdi_horas_decimal": round(pdi_min / 60, 2),
        "pdi_meta_minima": 20,
    }


def meses_disponiveis():
    """Retorna lista de meses (YYYY-MM) que têm registros, ordenados do mais recente
    para o mais antigo. Sempre inclui o mês atual no topo."""
    rows = base_query_for_user().with_entities(RDARecord.data).distinct().all()
    meses = set()
    for (data,) in rows:
        if data and len(data) >= 7:
            meses.add(data[:7])
    meses.add(datetime.utcnow().strftime("%Y-%m"))
    return sorted(meses, reverse=True)


def clientes_disponiveis_tap():
    """Retorna rótulos 'CTRL - Cliente' das TAPs com status Fechado.
    Concluído/Perdido/Aguardando não aparecem."""
    taps = TAP.query.filter_by(status_proposta="Fechado").order_by(TAP.ctrl_numero.asc()).all()
    return [t.rotulo_cliente for t in taps]


def entregaveis_por_projeto():
    """Mapa {rotulo_cliente: [entregaveis com valor_total > 0]} dos TAPs Fechados.
    Usado para filtrar a caixa 'O que foi realizado' por projeto."""
    taps = TAP.query.filter_by(status_proposta="Fechado").all()
    mapa = {}
    for t in taps:
        itens_validos = [i.entregavel for i in t.itens if i.valor_total > 0]
        mapa[t.rotulo_cliente] = itens_validos
    return mapa


def _score_janela(valor, min_ideal, max_ideal):
    """Retorna 100 dentro da janela ideal; decresce linearmente fora dela."""
    if valor is None:
        return 0
    if min_ideal <= valor <= max_ideal:
        return 100
    if valor < min_ideal:
        return max(0, valor / min_ideal * 100) if min_ideal > 0 else 0
    # valor > max_ideal: perde 1 ponto por % acima, até zerar
    excesso = valor - max_ideal
    return max(0, 100 - excesso)


def produtividade_colaboradores(mes_str: str = None):
    """Produtividade por colaborador no mês escolhido.

    Retorna (linhas, mes_key). Cada linha contém:
    - colaborador, total_horas, pdi_horas, projeto_horas
    - pct_aderencia (sobre 176h), pct_pdi, eficiencia (ponderada)
    - atividades_total, atividades_concluidas, pct_concluidas
    - total_registros, atrasados, pct_atrasos
    """
    hoje = datetime.utcnow()
    if mes_str and len(mes_str) == 7 and mes_str[4] == "-":
        try:
            datetime.strptime(mes_str, "%Y-%m")
            prefixo = mes_str
        except ValueError:
            prefixo = hoje.strftime("%Y-%m")
    else:
        prefixo = hoje.strftime("%Y-%m")

    records = RDARecord.query.filter(RDARecord.data.like(f"{prefixo}-%")).all()

    # Horas orçadas no TAP por (rotulo_cliente, entregavel)
    horas_previstas_map = {}
    for tap in TAP.query.all():
        rotulo = tap.rotulo_cliente
        for item in tap.itens:
            horas = (item.qtd_recursos or 0) * (item.tempo or 0)
            if horas > 0:
                horas_previstas_map[(rotulo, item.entregavel)] = horas

    # Para a eficiência: horas acumuladas (todo o histórico, não só o mês)
    # por grupo e por (colaborador, grupo). Atividade concluída pode consumir
    # horas em múltiplos meses; compará-la só com o mês corrente subestima.
    horas_grupo_all = defaultdict(float)
    horas_colab_grupo_all = defaultdict(float)  # (nome, grupo) -> horas
    for r in RDARecord.query.all():
        if (r.realizado or "") == "PDI":
            continue
        chave = (r.cliente or "", r.realizado or "")
        h = _minutos_registro(r) / 60
        horas_grupo_all[chave] += h
        nome_r = (r.colaborador or "").strip()
        if nome_r:
            horas_colab_grupo_all[(nome_r, chave)] += h

    # Mapa de prazo (fim_atividade) por grupo — para calcular SLA
    fim_atividade_map = _datas_previstas_por_atividade()

    colabs = defaultdict(lambda: {
        "total_min": 0,
        "pdi_min": 0,
        "projeto_min": 0,
        "horas_colab_por_grupo": defaultdict(float),
        "concluidas_set": set(),
        "atividades_set": set(),
        "total_registros": 0,
        "atrasados": 0,
        "data_conclusao_grupo": {},  # (cliente, ativ) -> max(data) dos Concluídos do colab
    })

    for r in records:
        nome = (r.colaborador or "").strip()
        if not nome:
            continue
        c = colabs[nome]
        m = _minutos_registro(r)
        c["total_min"] += m
        c["total_registros"] += 1
        if r.foi_atrasado:
            c["atrasados"] += 1
        if (r.realizado or "") == "PDI":
            c["pdi_min"] += m
        else:
            c["projeto_min"] += m
            chave = (r.cliente or "", r.realizado or "")
            c["horas_colab_por_grupo"][chave] += m / 60
            c["atividades_set"].add(chave)
            if r.status_rda == "Concluído":
                c["concluidas_set"].add(chave)
                atual = c["data_conclusao_grupo"].get(chave)
                if r.data and (atual is None or r.data > atual):
                    c["data_conclusao_grupo"][chave] = r.data

    resultado = []
    for nome, d in colabs.items():
        total_h = d["total_min"] / 60
        pdi_h = d["pdi_min"] / 60
        projeto_h = d["projeto_min"] / 60
        pct_aderencia = (total_h / 176 * 100) if 176 > 0 else 0
        pct_pdi = (pdi_h / total_h * 100) if total_h > 0 else 0

        # Eficiência média ponderada — considera apenas atividades
        # CONCLUÍDAS pelo colaborador no mês (senão a métrica fica
        # enganosa: um entregável em andamento mostraria "eficiência"
        # altíssima apenas porque há orçamento sobrando).
        # Usa horas acumuladas (histórico completo) para a comparação,
        # pois uma atividade concluída pode ter consumido horas em
        # meses anteriores também.
        soma_previsto_colab = 0.0
        soma_colab = 0.0
        for grupo in d["concluidas_set"]:
            h_prev = horas_previstas_map.get(grupo)
            if h_prev is None:
                continue
            h_colab_total = horas_colab_grupo_all.get((nome, grupo), 0)
            h_grupo_total = horas_grupo_all.get(grupo, 0)
            if h_colab_total <= 0 or h_grupo_total <= 0:
                continue
            share = h_colab_total / h_grupo_total
            soma_previsto_colab += h_prev * share
            soma_colab += h_colab_total
        eficiencia = (soma_previsto_colab / soma_colab * 100) if soma_colab > 0 else None

        pct_concluidas = (len(d["concluidas_set"]) / len(d["atividades_set"]) * 100) if d["atividades_set"] else 0
        pct_atrasos = (d["atrasados"] / d["total_registros"] * 100) if d["total_registros"] > 0 else 0

        # SLA: % de atividades concluídas entregues dentro do fim_atividade
        sla_total = 0
        sla_no_prazo = 0
        for chave in d["concluidas_set"]:
            fim = fim_atividade_map.get(chave)
            if not fim:
                continue
            sla_total += 1
            data_conc = d["data_conclusao_grupo"].get(chave)
            if data_conc and data_conc <= fim:
                sla_no_prazo += 1
        pct_sla = (sla_no_prazo / sla_total * 100) if sla_total > 0 else None

        # Aderência de capacidade: horas de projeto ÷ (176 − 20 PDI) = 156h
        CAPACIDADE_PROJETO = 156
        META_PDI = 20
        pct_aderencia_projeto = (projeto_h / CAPACIDADE_PROJETO * 100) if CAPACIDADE_PROJETO > 0 else 0

        # Notas 0-100 de cada dimensão
        s_aderencia = _score_janela(pct_aderencia_projeto, 90, 110)
        s_pdi = min(100.0, (pdi_h / META_PDI * 100)) if META_PDI > 0 else 0
        s_eficiencia = _score_janela(eficiencia, 90, 110) if eficiencia is not None else 50
        s_sla = pct_sla if pct_sla is not None else 50
        s_conclusao = pct_concluidas

        score_composto = (
            0.25 * s_aderencia +
            0.15 * s_pdi +
            0.25 * s_eficiencia +
            0.25 * s_sla +
            0.10 * s_conclusao
        )

        resultado.append({
            "colaborador": nome,
            "total_horas": round(total_h, 2),
            "pdi_horas": round(pdi_h, 2),
            "projeto_horas": round(projeto_h, 2),
            "pct_aderencia": round(pct_aderencia, 1),
            "pct_aderencia_projeto": round(pct_aderencia_projeto, 1),
            "pct_pdi": round(pct_pdi, 1),
            "pdi_cumprido": pdi_h >= META_PDI,
            "eficiencia": round(eficiencia, 1) if eficiencia is not None else None,
            "atividades_total": len(d["atividades_set"]),
            "atividades_concluidas": len(d["concluidas_set"]),
            "pct_concluidas": round(pct_concluidas, 1),
            "pct_sla": round(pct_sla, 1) if pct_sla is not None else None,
            "sla_no_prazo": sla_no_prazo,
            "sla_total": sla_total,
            "total_registros": d["total_registros"],
            "atrasados": d["atrasados"],
            "pct_atrasos": round(pct_atrasos, 1),
            "score": round(score_composto, 1),
        })

    resultado.sort(key=lambda x: x["colaborador"].lower())
    return resultado, prefixo


def gantt_projetos():
    """Gera dados para Frappe Gantt por projeto (TAP).

    Retorna lista de dicts {tap_id, cliente, ctrl_numero, tasks} onde tasks é
    a lista formatada para Frappe Gantt: id, name, start, end, progress,
    custom_class (para cores)."""
    hoje_iso = datetime.utcnow().strftime("%Y-%m-%d")

    # Coleta min/max data e minutos por grupo (cliente, atividade)
    rda_por_grupo = defaultdict(lambda: {"min": None, "max": None, "concluido": False, "min_realizados": 0})
    for r in RDARecord.query.all():
        if (r.realizado or "") == "PDI":
            continue
        chave = (r.cliente or "", r.realizado or "")
        g = rda_por_grupo[chave]
        if g["min"] is None or (r.data and r.data < g["min"]):
            g["min"] = r.data
        if g["max"] is None or (r.data and r.data > g["max"]):
            g["max"] = r.data
        if r.status_rda == "Concluído":
            g["concluido"] = True
        g["min_realizados"] += _minutos_registro(r)

    projetos = []
    for tap in TAP.query.order_by(TAP.ctrl_numero.asc()).all():
        rotulo = tap.rotulo_cliente
        tasks = []
        for idx, item in enumerate(tap.itens):
            horas_prev = (item.qtd_recursos or 0) * (item.tempo or 0)
            if horas_prev <= 0:
                continue
            rda_info = rda_por_grupo.get((rotulo, item.entregavel),
                                         {"min": None, "max": None, "concluido": False, "min_realizados": 0})

            # Start/end: usa datas do TAP se existirem, senão cai para RDA
            start = item.inicio_atividade or rda_info["min"]
            end = item.fim_atividade or rda_info["max"] or start
            if not start or not end:
                continue
            if end < start:
                end = start

            # Status → classe de cor
            if rda_info["concluido"]:
                if item.fim_atividade and rda_info["max"] and rda_info["max"] > item.fim_atividade:
                    custom_class = "bar-concluido-atrasado"
                else:
                    custom_class = "bar-concluido"
            elif item.fim_atividade and hoje_iso > item.fim_atividade:
                custom_class = "bar-atrasado"
            else:
                custom_class = "bar-andamento"

            # Progresso: horas realizadas / previstas × 100 (capped em 100)
            horas_real = rda_info["min_realizados"] / 60.0
            progress = min(100, round((horas_real / horas_prev * 100) if horas_prev > 0 else 0))

            tasks.append({
                "id": f"t{tap.id}-{idx}",
                "name": item.entregavel,
                "start": start,
                "end": end,
                "progress": progress,
                "custom_class": custom_class,
                "planejado_inicio": item.inicio_atividade or "",
                "planejado_fim": item.fim_atividade or "",
                "executado_inicio": rda_info["min"] or "",
                "executado_fim": rda_info["max"] or "",
                "horas_previstas": round(horas_prev, 2),
                "horas_realizadas": round(horas_real, 2),
            })

        if not tasks:
            continue

        projetos.append({
            "tap_id": tap.id,
            "cliente": tap.cliente,
            "ctrl_numero": tap.ctrl_numero,
            "tasks": tasks,
        })
    return projetos


def _dias_uteis_entre(dt_ini, dt_fim):
    """Conta dias úteis (seg-sex) entre dt_ini e dt_fim (ambos inclusos)."""
    from datetime import timedelta as _td
    if dt_fim < dt_ini:
        return 0
    count = 0
    d = dt_ini
    while d <= dt_fim:
        if d.weekday() < 5:
            count += 1
        d += _td(days=1)
    return count


def _somar_dias_uteis(dt_base, n):
    """Retorna dt_base + n dias úteis (dt_base conta como 1º dia útil)."""
    from datetime import timedelta as _td
    if n <= 0:
        return dt_base
    restantes = n - 1
    d = dt_base
    while restantes > 0:
        d += _td(days=1)
        if d.weekday() < 5:
            restantes -= 1
    return d


def _data_conclusao_por_grupo():
    """Mapa {(cliente, atividade): max(data)} entre registros Concluídos.
    Representa a data em que a atividade foi registrada como concluída."""
    from collections import defaultdict
    mapa = defaultdict(lambda: None)
    rows = (
        db.session.query(RDARecord.cliente, RDARecord.realizado, RDARecord.data)
        .filter(RDARecord.status_rda == "Concluído")
        .all()
    )
    for cliente, realizado, data in rows:
        chave = (cliente or "", realizado or "")
        atual = mapa.get(chave)
        if atual is None or (data and data > atual):
            mapa[chave] = data
    return mapa


def performance_previsto_vs_realizado():
    """Compara horas previstas (TAP) vs horas realizadas (RDA) por (projeto + entregável).

    Retorna lista de dicts com:
    - cliente, ctrl_numero, entregavel
    - horas_previstas, horas_realizadas
    - pct_consumido, saldo_horas
    - estado: 'ok' (< 80%), 'alerta' (80-100%), 'risco' (> 100%)
    """
    # Cache RDA: soma de minutos por (cliente_rotulo, entregavel)
    rda_cache = defaultdict(int)
    for r in RDARecord.query.all():
        chave = (r.cliente or "", r.realizado or "")
        rda_cache[chave] += _minutos_registro(r)

    conclusao_map = _data_conclusao_por_grupo()

    linhas = []
    taps = TAP.query.order_by(TAP.ctrl_numero.asc()).all()
    for tap in taps:
        rotulo = tap.rotulo_cliente
        for item in tap.itens:
            qtd = item.qtd_recursos or 0
            tempo = item.tempo or 0
            if tempo <= 0 or qtd <= 0:
                continue
            # Previsto = cronograma (apenas tempo). Realizado é normalizado
            # de horas-pessoa para horas-cronograma dividindo por qtd_recursos
            # para manter a comparação coerente.
            horas_previstas = tempo
            min_realizados = rda_cache.get((rotulo, item.entregavel), 0)
            horas_realizadas = (min_realizados / 60.0) / qtd
            pct = (horas_realizadas / horas_previstas * 100) if horas_previstas > 0 else 0
            saldo = horas_previstas - horas_realizadas
            if pct < 80:
                estado = "ok"
            elif pct <= 100:
                estado = "alerta"
            else:
                estado = "risco"
            # Data de Conclusão (MAX data dos RDAs Concluídos do grupo)
            data_conclusao = conclusao_map.get((rotulo, item.entregavel))
            conclusao_estado = "pendente"
            if data_conclusao and item.fim_atividade:
                if data_conclusao <= item.fim_atividade:
                    conclusao_estado = "no_prazo"
                else:
                    conclusao_estado = "atrasado"
            elif data_conclusao:
                conclusao_estado = "concluida_sem_prazo"
            linhas.append({
                "tap_id": tap.id,
                "cliente": tap.cliente,
                "ctrl_numero": tap.ctrl_numero,
                "status_tap": tap.status_proposta,
                "entregavel": item.entregavel,
                "horas_previstas": round(horas_previstas, 2),
                "horas_realizadas": round(horas_realizadas, 2),
                "pct": round(pct, 1),
                "saldo": round(saldo, 2),
                "estado": estado,
                "inicio_atividade": item.inicio_atividade,
                "fim_atividade": item.fim_atividade,
                "data_conclusao": data_conclusao,
                "conclusao_estado": conclusao_estado,
            })
    return linhas


def colaboradores_disponiveis():
    """Lista de nomes para o filtro Colaborador do relatório.
    - Admin e gestor: todos os usuários cadastrados + colaboradores presentes
      em registros históricos.
    - Colaborador: apenas o próprio nome (RDA é individual)."""
    from app.models import User
    if current_user.role == "colaborador":
        return [current_user.display_name]
    nomes = set()
    for u in User.query.all():
        nome = u.display_name
        if nome:
            nomes.add(nome.strip())
    for (nome,) in db.session.query(RDARecord.colaborador).distinct().all():
        if nome and nome.strip():
            nomes.add(nome.strip())
    return sorted(nomes, key=lambda s: s.lower())


def _form_inicial_vazio():
    """Retorna dicionário com nome pré-preenchido do usuário."""
    return {
        "id": "",
        "colaborador": current_user.display_name,
        "cliente": "",
        "data": "",
        "hora_inicio_manha": "",
        "hora_final_manha": "",
        "hora_inicio_tarde": "",
        "hora_final_tarde": "",
        "realizado": "",
        "status_rda": "Em Andamento",
    }


@main_bp.route("/")
@login_required
def index():
    _aplicar_auto_atraso()
    query = base_query_for_user()
    query, filters = aplicar_filtros(query)
    records = query.order_by(RDARecord.data.desc(), RDARecord.hora_inicio_manha.desc(), RDARecord.id.desc()).all()

    mes_dashboard = request.args.get("mes", "").strip()
    return render_template(
        "main/dashboard.html",
        records=records,
        total_registros=len(records),
        total_horas=horas_totais(records),
        filters=filters,
        form=_form_inicial_vazio(),
        editing=False,
        stats=stats_mes_atual(mes_dashboard),
        meses_disponiveis=meses_disponiveis(),
        formatar_data_br=formatar_data_br,
        clientes_tap=clientes_disponiveis_tap(),
        entregaveis_map=entregaveis_por_projeto(),
        entregaveis_default=TAP_ENTREGAVEIS,
        colaboradores_tap=colaboradores_disponiveis(),
    )


@main_bp.route("/save", methods=["POST"])
@login_required
def save_record():
    record_id = request.form.get("id", "").strip()

    payload = {
        "colaborador": request.form.get("colaborador", "").strip(),
        "cliente": request.form.get("cliente", "").strip(),
        "data": request.form.get("data", "").strip(),
        "hora_inicio_manha": request.form.get("hora_inicio_manha", "").strip(),
        "hora_final_manha": request.form.get("hora_final_manha", "").strip(),
        "hora_inicio_tarde": request.form.get("hora_inicio_tarde", "").strip(),
        "hora_final_tarde": request.form.get("hora_final_tarde", "").strip(),
        "realizado": request.form.get("realizado", "").strip(),
        "status_rda": request.form.get("status_rda", "Em Andamento").strip(),
    }

    try:
        obrigatorios = [payload["colaborador"], payload["cliente"], payload["data"], payload["realizado"]]
        if not all(obrigatorios):
            raise ValueError("Preencha Colaborador, Cliente, Data e o que foi realizado.")

        # Campos de hora em branco viram "00:00"
        for campo in ["hora_inicio_manha", "hora_final_manha", "hora_inicio_tarde", "hora_final_tarde"]:
            if not payload[campo]:
                payload[campo] = "00:00"

        # Regra: Início da Tarde >= Final da Manhã + 1h (só quando ambos preenchidos)
        fm_min = _minutos_intervalo("00:00", payload["hora_final_manha"])
        it_min = _minutos_intervalo("00:00", payload["hora_inicio_tarde"])
        if fm_min > 0 and it_min > 0 and it_min < fm_min + 60:
            raise ValueError(
                "Início do período da Tarde inválido. A Hora Início da Tarde "
                "deve ser no mínimo 1 hora após a Hora Final da Manhã. "
                "Corrija o campo 'Hora Início' do período da tarde."
            )

        payload["duracao"] = calcular_duracao_total(
            payload["hora_inicio_manha"], payload["hora_final_manha"],
            payload["hora_inicio_tarde"], payload["hora_final_tarde"],
        )

        if payload["duracao"] == "00:00":
            raise ValueError("Informe pelo menos um período (manhã ou tarde).")

        if record_id:
            record = RDARecord.query.filter_by(id=int(record_id)).first_or_404()
            if current_user.role == "colaborador" and record.user_id != current_user.id:
                raise ValueError("Você não tem permissão para editar este registro.")
            for key, value in payload.items():
                setattr(record, key, value)
        else:
            record = RDARecord(user_id=current_user.id, **payload)
            db.session.add(record)

        # Regra: se o grupo (cliente + atividade) já tem algum registro
        # Concluído, este registro também passa automaticamente para Concluído
        # (exceto PDI, que é atividade livre). Aplica-se antes da flag sticky.
        if (
            payload["realizado"] != "PDI"
            and payload["cliente"]
            and payload["realizado"]
            and payload["status_rda"] == "Em Andamento"
        ):
            ja_concluido = (
                RDARecord.query.filter(
                    RDARecord.cliente == payload["cliente"],
                    RDARecord.realizado == payload["realizado"],
                    RDARecord.status_rda == "Concluído",
                    RDARecord.id != record.id if record.id else True,
                )
                .first()
            )
            if ja_concluido is not None:
                payload["status_rda"] = "Concluído"
                record.status_rda = "Concluído"

        # Flag sticky: uma vez Atrasado, permanece contabilizado em Atrasado
        # mesmo após conclusão.
        if payload["status_rda"] == "Atrasado":
            record.foi_atrasado = True

        # Auto-atraso por fim_atividade do TAP: considera hoje > fim OU a
        # data do próprio registro > fim (trabalho após o prazo).
        hoje_iso = datetime.utcnow().strftime("%Y-%m-%d")
        mapa = _datas_previstas_por_atividade()
        if _esta_atrasado_por_data(
            payload["cliente"], payload["realizado"],
            hoje_iso, mapa, data_registro=payload["data"],
        ):
            record.foi_atrasado = True
            if payload["status_rda"] != "Concluído":
                record.status_rda = "Atrasado"
                payload["status_rda"] = "Atrasado"

        # Cascata forward: se esta atividade foi salva como Concluída, todos
        # os outros registros da mesma (cliente + atividade) em Em Andamento
        # passam para Concluído. Registros Atrasado permanecem Atrasado
        # (regra sticky). PDI é ignorado.
        cascata = 0
        if (
            payload["status_rda"] == "Concluído"
            and payload["realizado"] != "PDI"
            and payload["cliente"]
            and payload["realizado"]
        ):
            cascata = (
                RDARecord.query.filter(
                    RDARecord.cliente == payload["cliente"],
                    RDARecord.realizado == payload["realizado"],
                    RDARecord.status_rda == "Em Andamento",
                )
                .update({"status_rda": "Concluído"}, synchronize_session=False)
            )

        db.session.commit()
        _registrar_audit("editar" if record_id else "criar", record)
        if cascata:
            flash(
                f"Registro salvo. {cascata} outro(s) registro(s) da mesma atividade "
                f"foram marcados como Concluído.",
                "success",
            )
        else:
            flash("Registro salvo com sucesso.", "success")
        return redirect(url_for("main.index"))

    except ValueError as exc:
        flash(str(exc), "error")
        query = base_query_for_user()
        query, filters = aplicar_filtros(query)
        records = query.order_by(RDARecord.data.desc(), RDARecord.hora_inicio_manha.desc(), RDARecord.id.desc()).all()
        form_data = {"id": record_id, **payload}
        return render_template(
            "main/dashboard.html",
            records=records,
            total_registros=len(records),
            total_horas=horas_totais(records),
            filters=filters,
            form=form_data,
            editing=bool(record_id),
            stats=stats_mes_atual(),
            formatar_data_br=formatar_data_br,
            clientes_tap=clientes_disponiveis_tap(),
            entregaveis_map=entregaveis_por_projeto(),
            entregaveis_default=TAP_ENTREGAVEIS,
        )


@main_bp.route("/edit/<int:record_id>")
@login_required
def edit_record(record_id: int):
    record = RDARecord.query.filter_by(id=record_id).first_or_404()
    if current_user.role == "colaborador" and record.user_id != current_user.id:
        flash("Você não tem permissão para editar este registro.", "error")
        return redirect(url_for("main.index"))

    query = base_query_for_user()
    query, filters = aplicar_filtros(query)
    records = query.order_by(RDARecord.data.desc(), RDARecord.hora_inicio_manha.desc(), RDARecord.id.desc()).all()

    return render_template(
        "main/dashboard.html",
        records=records,
        total_registros=len(records),
        total_horas=horas_totais(records),
        filters=filters,
        form=record,
        editing=True,
        stats=stats_mes_atual(request.args.get("mes", "").strip()),
        meses_disponiveis=meses_disponiveis(),
        formatar_data_br=formatar_data_br,
        clientes_tap=clientes_disponiveis_tap(),
        entregaveis_map=entregaveis_por_projeto(),
        entregaveis_default=TAP_ENTREGAVEIS,
        colaboradores_tap=colaboradores_disponiveis(),
    )


@main_bp.route("/delete/<int:record_id>", methods=["POST"])
@login_required
def delete_record(record_id: int):
    record = RDARecord.query.filter_by(id=record_id).first_or_404()
    if current_user.role == "colaborador" and record.user_id != current_user.id:
        flash("Você não tem permissão para excluir este registro.", "error")
        return redirect(url_for("main.index"))

    # Salva snapshot antes de deletar para o log
    snapshot = {
        "id": record.id,
        "cliente": record.cliente,
        "realizado": record.realizado,
        "data": record.data,
    }
    db.session.delete(record)
    db.session.commit()

    class _Snap:
        pass
    snap = _Snap()
    snap.id = snapshot["id"]
    snap.cliente = snapshot["cliente"]
    snap.realizado = snapshot["realizado"]
    snap.data = snapshot["data"]
    _registrar_audit("excluir", snap, record_id=snapshot["id"])

    flash("Registro excluído.", "success")
    return redirect(url_for("main.index"))


@main_bp.route("/performance")
@login_required
def performance():
    # Acesso liberado para todos os usuários autenticados — transparência da
    # performance real dos projetos para toda a equipe.
    _aplicar_auto_atraso()
    linhas = performance_previsto_vs_realizado()

    f_cliente = request.args.get("f_cliente", "").strip()
    f_estado = request.args.get("f_estado", "").strip()
    if f_cliente:
        linhas = [l for l in linhas if f_cliente.lower() in l["cliente"].lower()]
    if f_estado in ("ok", "alerta", "risco"):
        linhas = [l for l in linhas if l["estado"] == f_estado]

    # Agrupa por projeto (cliente + ctrl_numero) para subtotais
    grupos = {}
    for l in linhas:
        chave = (l["cliente"], l["ctrl_numero"], l["status_tap"], l["tap_id"])
        grupos.setdefault(chave, []).append(l)

    projetos = []
    for (cliente, ctrl, status_tap, tap_id), itens in grupos.items():
        prev = sum(i["horas_previstas"] for i in itens)
        real = sum(i["horas_realizadas"] for i in itens)
        pct = (real / prev * 100) if prev > 0 else 0
        if pct < 80:
            estado = "ok"
        elif pct <= 100:
            estado = "alerta"
        else:
            estado = "risco"
        projetos.append({
            "cliente": cliente,
            "ctrl_numero": ctrl,
            "status_tap": status_tap,
            "tap_id": tap_id,
            "itens": itens,
            "horas_previstas": round(prev, 2),
            "horas_realizadas": round(real, 2),
            "pct": round(pct, 1),
            "saldo": round(prev - real, 2),
            "estado": estado,
            "qtd_risco": sum(1 for i in itens if i["estado"] == "risco"),
        })
    projetos.sort(key=lambda p: p["ctrl_numero"])

    # Totais gerais
    total_prev = sum(p["horas_previstas"] for p in projetos)
    total_real = sum(p["horas_realizadas"] for p in projetos)
    total_pct = (total_real / total_prev * 100) if total_prev > 0 else 0
    total_risco = sum(1 for p in projetos if p["estado"] == "risco")

    clientes_filtro = sorted({p["cliente"] for p in projetos}, key=lambda s: s.lower())

    return render_template(
        "main/performance.html",
        projetos=projetos,
        filtros={"f_cliente": f_cliente, "f_estado": f_estado},
        clientes_filtro=clientes_filtro,
        total_prev=round(total_prev, 2),
        total_real=round(total_real, 2),
        total_pct=round(total_pct, 1),
        total_risco=total_risco,
        total_projetos=len(projetos),
    )


@main_bp.route("/performance/gantt")
@login_required
def performance_gantt():
    _aplicar_auto_atraso()
    projetos = gantt_projetos()
    return render_template("main/gantt.html", projetos=projetos)


# ============================================================================
# Predição (modelo de capacidade ex-ante)
# ----------------------------------------------------------------------------
# Compara o escopo planejado (HH = qtd_recursos × tempo) com a capacidade útil
# da janela planejada (recursos × dias_úteis × 8h). Responde "o projeto cabe
# nesse prazo com esses recursos?", independentemente do que já foi lançado.
# Parâmetros em PREDICAO_HORAS_DIA e PREDICAO_LIMITE_NO_LIMITE (ver Premissas).
# ============================================================================

PREDICAO_HORAS_DIA = 8
PREDICAO_LIMITE_NO_LIMITE = 0.95


def predicao_capacidade():
    """Modelo de capacidade por entregável + roll-up por projeto.

    Retorna (entregaveis, projetos) onde:
      entregaveis = lista de dicts por TAPItem (completo, parcial ou sem escopo)
      projetos    = lista de dicts por TAP com agregação e risco
    """
    from math import ceil

    def _parse(iso):
        if not iso:
            return None
        try:
            return datetime.strptime(iso, "%Y-%m-%d").date()
        except (ValueError, TypeError):
            return None

    entregaveis = []
    projetos = []
    for tap in TAP.query.order_by(TAP.ctrl_numero.asc()).all():
        hh_valor = tap.hh_valor or 0
        itens_proj = []
        primeiro_ini_plan = None
        ultimo_fim_plan = None
        primeiro_ini_prev = None
        ultimo_fim_prev = None
        cont_atraso = 0
        cont_no_limite = 0
        cont_com_folga = 0
        cont_base_parcial = 0
        cont_base_completa = 0
        cont_ativos = 0
        soma_hh_extras = 0.0
        soma_valor_corrigido = 0.0
        soma_custo_adicional = 0.0
        desvios_completos = []

        for item in tap.itens:
            qtd = item.qtd_recursos or 0
            tempo = item.tempo or 0
            pc = item.percentual_correcao or 0
            hh_escopo = qtd * tempo
            valor_total = hh_escopo * hh_valor
            valor_corrigido = valor_total * (1 + pc / 100.0)
            ini = _parse(item.inicio_atividade)
            fim = _parse(item.fim_atividade)

            base = {
                "tap_id": tap.id,
                "ctrl_numero": tap.ctrl_numero,
                "cliente": tap.cliente,
                "status_tap": tap.status_proposta,
                "ordem": item.ordem,
                "entregavel": item.entregavel,
                "qtd_recursos": qtd,
                "tempo": tempo,
                "percentual_correcao": pc,
                "inicio_planejado": item.inicio_atividade or "",
                "fim_planejado": item.fim_atividade or "",
                "hh_escopo": round(hh_escopo, 2),
                "valor_corrigido": round(valor_corrigido, 2),
                "dias_uteis_planejados": None,
                "capacidade_hh": None,
                "indice_carga": None,
                "recursos_minimos": None,
                "inicio_previsto": item.inicio_atividade or "",
                "fim_previsto": "",
                "desvio_dias": None,
                "hh_extras": None,
                "custo_adicional": None,
                "custo_total_mantendo_prazo": round(valor_corrigido, 2) if valor_corrigido else 0,
                "status_prazo": "Não planejado",
                "flag_dados": "Sem escopo",
            }

            if qtd <= 0 or tempo <= 0:
                entregaveis.append(base)
                continue

            cont_ativos += 1

            if ini and fim:
                dias_plan = _dias_uteis_entre(ini, fim)
                capacidade = qtd * dias_plan * PREDICAO_HORAS_DIA
                indice = hh_escopo / capacidade if capacidade > 0 else None
                recursos_min = ceil(hh_escopo / (dias_plan * PREDICAO_HORAS_DIA)) if dias_plan > 0 else None
                dias_necess = ceil(hh_escopo / (qtd * PREDICAO_HORAS_DIA)) if qtd > 0 else 0
                fim_previsto = _somar_dias_uteis(ini, dias_necess)
                desvio = _dias_uteis_entre(fim, fim_previsto) - 1 if fim_previsto >= fim else -(_dias_uteis_entre(fim_previsto, fim) - 1)
                hh_extras = max(0.0, hh_escopo - capacidade)
                custo_ad = hh_extras * hh_valor

                if indice is None:
                    status = "Não planejado"
                elif indice > 1.0:
                    status = "Atraso previsto"
                    cont_atraso += 1
                elif indice >= PREDICAO_LIMITE_NO_LIMITE:
                    status = "No limite"
                    cont_no_limite += 1
                else:
                    status = "Com folga"
                    cont_com_folga += 1

                cont_base_completa += 1
                desvios_completos.append(desvio)
                base.update({
                    "dias_uteis_planejados": dias_plan,
                    "capacidade_hh": round(capacidade, 2),
                    "indice_carga": round(indice, 4) if indice is not None else None,
                    "recursos_minimos": recursos_min,
                    "inicio_previsto": ini.strftime("%Y-%m-%d"),
                    "fim_previsto": fim_previsto.strftime("%Y-%m-%d"),
                    "desvio_dias": desvio,
                    "hh_extras": round(hh_extras, 2),
                    "custo_adicional": round(custo_ad, 2),
                    "custo_total_mantendo_prazo": round(valor_corrigido + custo_ad, 2),
                    "status_prazo": status,
                    "flag_dados": "Completo",
                })
                soma_hh_extras += hh_extras
                soma_custo_adicional += custo_ad

                if primeiro_ini_plan is None or ini < primeiro_ini_plan:
                    primeiro_ini_plan = ini
                if ultimo_fim_plan is None or fim > ultimo_fim_plan:
                    ultimo_fim_plan = fim
                if primeiro_ini_prev is None or ini < primeiro_ini_prev:
                    primeiro_ini_prev = ini
                if ultimo_fim_prev is None or fim_previsto > ultimo_fim_prev:
                    ultimo_fim_prev = fim_previsto
            else:
                cont_base_parcial += 1
                base.update({
                    "status_prazo": "Base parcial",
                    "flag_dados": "Data incompleta",
                })

            soma_valor_corrigido += valor_corrigido
            entregaveis.append(base)
            itens_proj.append(base)

        if cont_ativos == 0:
            continue

        if cont_base_parcial > 0:
            desvio_projeto = None
        else:
            desvio_projeto = max(desvios_completos) if desvios_completos else 0

        if cont_base_parcial >= 2 or cont_atraso >= 2:
            risco = "Crítico"
            leitura = "Replanejar agora: há atraso previsto e lacunas de cronograma."
        elif cont_atraso >= 1 or cont_base_parcial >= 1:
            risco = "Alto"
            leitura = "Ajustar prazo ou fechar datas faltantes."
        elif cont_no_limite >= 1:
            risco = "Moderado"
            leitura = "Projeto viável, mas sem gordura de cronograma."
        else:
            risco = "Baixo"
            leitura = "Cronograma com folga e recursos suficientes."

        projetos.append({
            "tap_id": tap.id,
            "ctrl_numero": tap.ctrl_numero,
            "cliente": tap.cliente,
            "status_tap": tap.status_proposta,
            "entregaveis_ativos": cont_ativos,
            "base_completa": cont_base_completa,
            "base_parcial": cont_base_parcial,
            "atraso_previsto": cont_atraso,
            "no_limite": cont_no_limite,
            "com_folga": cont_com_folga,
            "hh_extras": round(soma_hh_extras, 2),
            "valor_corrigido_total": round(soma_valor_corrigido, 2),
            "custo_adicional": round(soma_custo_adicional, 2),
            "custo_total_mantendo_prazo": round(soma_valor_corrigido + soma_custo_adicional, 2),
            "primeiro_inicio_planejado": primeiro_ini_plan.strftime("%Y-%m-%d") if primeiro_ini_plan else "",
            "ultimo_fim_planejado": ultimo_fim_plan.strftime("%Y-%m-%d") if ultimo_fim_plan else "",
            "primeiro_inicio_previsto": primeiro_ini_prev.strftime("%Y-%m-%d") if primeiro_ini_prev else "",
            "ultimo_fim_previsto": ultimo_fim_prev.strftime("%Y-%m-%d") if ultimo_fim_prev else "",
            "desvio_dias": desvio_projeto,
            "risco": risco,
            "leitura": leitura,
            "itens": itens_proj,
        })

    projetos.sort(key=lambda p: p["ctrl_numero"])
    return entregaveis, projetos


@main_bp.route("/predicao")
@login_required
def predicao():
    _, projetos = predicao_capacidade()
    f_projeto = request.args.get("f_projeto", "").strip()
    todos = list(projetos)
    if f_projeto:
        try:
            tap_id = int(f_projeto)
            projetos = [p for p in projetos if p["tap_id"] == tap_id]
        except ValueError:
            pass
    premissas = {
        "horas_uteis_dia": PREDICAO_HORAS_DIA,
        "limite_no_limite": PREDICAO_LIMITE_NO_LIMITE,
    }
    return render_template(
        "main/predicao.html",
        projetos=projetos,
        todos=todos,
        filtro_projeto=f_projeto,
        premissas=premissas,
    )


@main_bp.route("/avaliacao")
@login_required
def avaliacao():
    if not current_user.is_admin():
        abort(403)
    entregaveis, projetos = predicao_capacidade()
    total_itens = len(entregaveis)
    base_completa = sum(1 for e in entregaveis if e.get("base_completa"))
    META_AMOSTRAS = 50
    fechados = 0
    progresso_pct = round(min(100.0, (fechados / META_AMOSTRAS) * 100), 1) if META_AMOSTRAS else 0.0
    return render_template(
        "main/avaliacao.html",
        total_itens=total_itens,
        base_completa=base_completa,
        total_projetos=len(projetos),
        fechados=fechados,
        meta=META_AMOSTRAS,
        progresso_pct=progresso_pct,
    )


@main_bp.route("/predicao/export/xlsx")
@login_required
def predicao_export_xlsx():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    entregaveis, projetos = predicao_capacidade()

    wb = Workbook()

    # Aba 1 — EAP (fonte)
    ws1 = wb.active
    ws1.title = "EAP"
    ws1.append(["CTRL Nº", "Cliente", "Status Proposta", "HH (R$/h)", "Ordem",
                "Entregável", "Qtd Recursos", "Tempo (h)", "% Correção",
                "Início Atividade", "Fim Atividade", "Valor Total (R$)",
                "Valor Corrigido (R$)"])
    for tap in TAP.query.order_by(TAP.ctrl_numero.asc()).all():
        hh = tap.hh_valor or 0
        for item in tap.itens:
            qtd = item.qtd_recursos or 0
            tempo = item.tempo or 0
            pc = item.percentual_correcao or 0
            vt = qtd * tempo * hh
            vc = vt * (1 + pc / 100.0)
            ws1.append([tap.ctrl_numero, tap.cliente, tap.status_proposta, hh, item.ordem,
                        item.entregavel, qtd, tempo, pc,
                        item.inicio_atividade or "", item.fim_atividade or "",
                        round(vt, 2), round(vc, 2)])

    # Aba 2 — Predicao_Entregaveis
    ws2 = wb.create_sheet("Predicao_Entregaveis")
    ws2.append(["CTRL Nº", "Cliente", "Status Proposta", "Ordem", "Entregável",
                "Qtd Recursos", "Tempo (h)", "Início Planejado", "Fim Planejado",
                "Valor Corrigido (R$)", "HH Total Escopo", "Dias Úteis Planejados",
                "Capacidade HH no Prazo", "Índice de Carga",
                "Recursos Mínimos no Prazo", "Início Previsto", "Fim Previsto",
                "Desvio Prazo (dias úteis)", "HH Extras p/ Cumprir Prazo",
                "Custo Adicional p/ Cumprir Prazo (R$)",
                "Custo Total se Mantiver Prazo (R$)", "Status Prazo", "Flag Dados"])
    for e in entregaveis:
        ws2.append([e["ctrl_numero"], e["cliente"], e["status_tap"], e["ordem"],
                    e["entregavel"], e["qtd_recursos"], e["tempo"],
                    e["inicio_planejado"], e["fim_planejado"], e["valor_corrigido"],
                    e["hh_escopo"], e["dias_uteis_planejados"], e["capacidade_hh"],
                    e["indice_carga"], e["recursos_minimos"], e["inicio_previsto"],
                    e["fim_previsto"], e["desvio_dias"], e["hh_extras"],
                    e["custo_adicional"], e["custo_total_mantendo_prazo"],
                    e["status_prazo"], e["flag_dados"]])

    # Aba 3 — Predicao_Projetos
    ws3 = wb.create_sheet("Predicao_Projetos")
    ws3.append(["CTRL Nº", "Cliente", "Status Proposta", "Entregáveis Ativos",
                "Base Completa", "Base Parcial", "Atraso Previsto", "No Limite",
                "HH Extras p/ Cumprir Prazo", "Valor Corrigido Total (R$)",
                "Custo Adicional p/ Cumprir Prazo (R$)",
                "Custo Total se Mantiver Prazo (R$)", "Primeiro Início Planejado",
                "Último Fim Planejado", "Primeiro Início Previsto",
                "Último Fim Previsto", "Desvio do Projeto (dias úteis)",
                "Risco do Projeto", "Leitura Executiva"])
    for p in projetos:
        ws3.append([p["ctrl_numero"], p["cliente"], p["status_tap"],
                    p["entregaveis_ativos"], p["base_completa"], p["base_parcial"],
                    p["atraso_previsto"], p["no_limite"], p["hh_extras"],
                    p["valor_corrigido_total"], p["custo_adicional"],
                    p["custo_total_mantendo_prazo"], p["primeiro_inicio_planejado"],
                    p["ultimo_fim_planejado"], p["primeiro_inicio_previsto"],
                    p["ultimo_fim_previsto"], p["desvio_dias"], p["risco"],
                    p["leitura"]])

    # Aba 4 — Premissas
    ws4 = wb.create_sheet("Premissas")
    ws4.append(["Parâmetro", "Valor", "Observação"])
    ws4.append(["Horas úteis por dia", PREDICAO_HORAS_DIA,
                "Base para transformar horas em duração útil."])
    ws4.append(["Limite para status 'No limite'", PREDICAO_LIMITE_NO_LIMITE,
                "Índice acima disso já considera a carga apertada."])

    # Estilização mínima — cabeçalhos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0F4C81", end_color="0F4C81", fill_type="solid")
    for ws in (ws1, ws2, ws3, ws4):
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 30
        for col_idx in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    stamp = datetime.now().strftime("%Y%m%d_%H%M")
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"predicao_capacidade_{stamp}.xlsx",
    )


@main_bp.route("/performance/colaboradores")
@login_required
def performance_colaboradores():
    if not (current_user.is_admin() or getattr(current_user, "acesso_produtividade", False)):
        flash("Acesso restrito — peça autorização a um administrador.", "error")
        return redirect(url_for("main.index"))

    _aplicar_auto_atraso()
    mes = request.args.get("mes", "").strip()
    linhas, mes_key = produtividade_colaboradores(mes)

    try:
        mes_label = datetime.strptime(mes_key, "%Y-%m").strftime("%m/%Y")
    except ValueError:
        mes_label = datetime.utcnow().strftime("%m/%Y")

    # Benchmark da equipe
    media = {}
    if linhas:
        n = len(linhas)
        media = {
            "score": round(sum(l["score"] for l in linhas) / n, 1),
            "aderencia": round(sum(l["pct_aderencia_projeto"] for l in linhas) / n, 1),
            "pdi": round(sum(l["pct_pdi"] for l in linhas) / n, 1),
            "conclusao": round(sum(l["pct_concluidas"] for l in linhas) / n, 1),
        }

    return render_template(
        "main/produtividade.html",
        linhas=linhas,
        mes_key=mes_key,
        mes_label=mes_label,
        meses_disponiveis=meses_disponiveis(),
        media=media,
    )


@main_bp.route("/api/records")
@login_required
def api_records():
    query = base_query_for_user()
    query, filters = aplicar_filtros(query)
    records = query.order_by(RDARecord.data.desc(), RDARecord.hora_inicio_manha.desc(), RDARecord.id.desc()).all()

    return jsonify({
        "user": current_user.username,
        "role": current_user.role,
        "filters": filters,
        "total_registros": len(records),
        "total_horas": horas_totais(records),
        "records": [
            {
                "id": r.id,
                "colaborador": r.colaborador,
                "cliente": r.cliente,
                "data": formatar_data_br(r.data),
                "hora_inicio_manha": r.hora_inicio_manha,
                "hora_final_manha": r.hora_final_manha,
                "hora_inicio_tarde": r.hora_inicio_tarde,
                "hora_final_tarde": r.hora_final_tarde,
                "duracao": r.duracao,
                "realizado": r.realizado,
                "status_rda": r.status_rda,
                "owner_username": r.owner.username if r.owner else None,
            }
            for r in records
        ],
    })


@main_bp.route("/export/csv")
@login_required
def export_csv():
    query = base_query_for_user()
    query, _ = aplicar_filtros(query)
    records = query.order_by(RDARecord.data.desc(), RDARecord.hora_inicio_manha.desc(), RDARecord.id.desc()).all()

    output = StringIO()
    writer = csv.writer(output, delimiter=";")
    writer.writerow([
        "Colaborador", "Cliente", "Data",
        "Manhã Início", "Manhã Final", "Tarde Início", "Tarde Final",
        "Duração", "O que foi Realizado", "Status", "Dono do Registro",
    ])

    for item in records:
        writer.writerow([
            item.colaborador,
            item.cliente,
            formatar_data_br(item.data),
            item.hora_inicio_manha,
            item.hora_final_manha,
            item.hora_inicio_tarde,
            item.hora_final_tarde,
            item.duracao,
            item.realizado,
            item.status_rda,
            item.owner.username if item.owner else "",
        ])

    response = make_response("\ufeff" + output.getvalue())
    response.headers["Content-Type"] = "text/csv; charset=utf-8"
    response.headers["Content-Disposition"] = "attachment; filename=rda_relatorio.csv"
    return response


@main_bp.route("/export/pdf")
@login_required
def export_pdf():
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    query = base_query_for_user()
    query, _ = aplicar_filtros(query)
    records = query.order_by(RDARecord.data.desc(), RDARecord.hora_inicio_manha.desc(), RDARecord.id.desc()).all()

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        leftMargin=10 * mm, rightMargin=10 * mm, topMargin=12 * mm, bottomMargin=12 * mm,
    )

    styles = getSampleStyleSheet()
    titulo = Paragraph("<b>Relatório Diário de Atividade (RDA)</b>", styles["Title"])
    gerado = Paragraph(
        f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')} — Total: {len(records)} registros — "
        f"Horas: {horas_totais(records)}",
        styles["Normal"],
    )

    cell_style = ParagraphStyle("cell", parent=styles["BodyText"], fontSize=7, leading=9)
    head_style = ParagraphStyle("head", parent=styles["BodyText"], fontSize=7, leading=9, textColor=colors.white, alignment=1)

    headers = ["Colaborador", "Cliente", "Data", "Manhã Ini", "Manhã Fim", "Tarde Ini", "Tarde Fim", "Duração", "Realizado", "Status"]
    data = [[Paragraph(h, head_style) for h in headers]]

    for r in records:
        data.append([
            Paragraph(r.colaborador or "", cell_style),
            Paragraph(r.cliente or "", cell_style),
            Paragraph(formatar_data_br(r.data), cell_style),
            Paragraph(r.hora_inicio_manha or "", cell_style),
            Paragraph(r.hora_final_manha or "", cell_style),
            Paragraph(r.hora_inicio_tarde or "", cell_style),
            Paragraph(r.hora_final_tarde or "", cell_style),
            Paragraph(r.duracao or "", cell_style),
            Paragraph(r.realizado or "", cell_style),
            Paragraph(r.status_rda or "", cell_style),
        ])

    col_widths = [30*mm, 28*mm, 18*mm, 16*mm, 16*mm, 16*mm, 16*mm, 16*mm, 70*mm, 22*mm]
    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2d5a")),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
    ]))

    doc.build([titulo, Spacer(1, 4 * mm), gerado, Spacer(1, 4 * mm), table])
    buffer.seek(0)

    return send_file(
        buffer, as_attachment=True,
        download_name=f"rda_relatorio_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf",
        mimetype="application/pdf",
    )
